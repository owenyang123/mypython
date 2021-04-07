#!/apollo/env/bin python

from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side
import argparse
    
class xlsx_styles:
    headers = ["a_hostname","a_interface","a_optic_type","a_connector_type","a_fiber_type","a_port_speed","z_fiber_type","z_connector_type","z_optic_type","z_port_speed","z_interface","z_hostname"]
    header_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='medium'))
    content_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    prefill_svc = ["a-host", "a-int", "SFP+ 10G LR","LC","smf","10G","smf","MTP Breakout to 4x LC","QSFP+ PSM4 LR","10G", "z-int", "z-host"]
    green_bgcolor = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type = "solid")
    font_size = Font(size=12)
    svc_16int = ["xe-0/0/0:0","xe-0/0/0:1","xe-0/0/0:2","xe-0/0/0:3","xe-0/0/8:0","xe-0/0/8:1","xe-0/0/8:2","xe-0/0/8:3","xe-0/0/16:0","xe-0/0/16:1","xe-0/0/16:2","xe-0/0/16:3","xe-0/0/24:0","xe-0/0/24:1","xe-0/0/24:2","xe-0/0/24:3"]
#replace it with actual port assignment for vc-svc 32 port assignment
    svc_32int = ["xyz"]
    
# initialize the cutsheet xlsx file, for VC-EDG<>VC-SVC cutsheet
def initialize_xlsx(ws_cutsheet, int_count, hostnames):
    
    ws_cutsheet.title = 'VC-EDG to VC-SVC'
    
    for i in range(len(xlsx_styles.headers)):
        ws_cutsheet.cell(row = 1, column = i + 1).value = xlsx_styles.headers[i]
        ws_cutsheet.cell(row = 1, column = i + 1).border = xlsx_styles.header_border
        ws_cutsheet.cell(row = 1, column = i + 1).font = xlsx_styles.font_size
    
        # Pre-populate data in cutsheet. Range: Row 2 to last device's last row.
        for blank_row in range(int_count * len(hostnames)):
            ws_cutsheet.cell(row = blank_row + 2, column = i + 1).value = xlsx_styles.prefill_svc[i]
            # set thin border lines
            ws_cutsheet.cell(row = blank_row + 2, column = i + 1).border = xlsx_styles.content_border
            # set background color to green
            ws_cutsheet.cell(row = blank_row + 2, column = i + 1).fill = xlsx_styles.green_bgcolor
            ws_cutsheet.cell(row = blank_row + 2, column = i + 1).font = xlsx_styles.font_size
    
# fill hostnames and interface ids in the cutsheet
def datafill_host_int(hostnames, int_count, avail_port_dict, ws_cutsheet,host_col_idx = 0, int_col_idx = 1):
    
    offset_row = int_count
    
    for i in range(int_count):
        for index in range(len(hostnames)):
            ws_cutsheet.cell(row = i + 2 + offset_row * index, column = host_col_idx + 1).value = hostnames[index]
            ws_cutsheet.cell(row = i + 2 + offset_row * index, column = int_col_idx + 1).value = avail_port_dict[hostnames[index]][i]
    
def svc_available_ports(int_count, hostnames):
    avail_port_dict = {}
    for hostname in hostnames:
        if int_count == 16:
            avail_port_dict[hostname] = xlsx_styles.svc_16int
        else:
            avail_port_dict[hostname] = xlsx_styles.svc_32int
    return avail_port_dict

def edg_available_ports(int_count, hostnames):
    '''
    return the availbe ports in vc-edg routers listed in hostnames
    '''
    avail_port_dict = {}
    available_intf_list = []
    for hostname in hostnames:
        avail_port_dict[hostname] = available_intf_list
    
'''
#Sample values
'''

print("Test")
int_count_z = 16
int_count_a = int_count_z
hostnames_a = ["icn54-vc-edg-r1","icn54-vc-edg-r2"]
hostnames_z = ["icn54-54-vc-svc-r311","icn54-54-vc-svc-r312"]
    
# avail_port_dict format: {'hostname':list of ports, 'hostname', list of ports}
avail_port_dict_a = {'icn54-vc-edg-r1': ['xe-0/1/3', 'xe-0/2/2', 'xe-0/2/3', 'xe-0/3/2', 'xe-0/3/3', 'xe-1/2/2', 'xe-1/3/1', 'xe-2/1/2', 'xe-2/1/3', 'xe-2/2/2', 'xe-2/2/3', 'xe-2/3/3', 'xe-3/0/2', 'xe-3/1/2', 'xe-3/2/1', 'xe-3/2/2'], 'icn54-vc-edg-r2': ['xe-0/1/3', 'xe-0/2/2', 'xe-0/2/3', 'xe-0/3/2', 'xe-0/3/3', 'xe-1/2/2', 'xe-1/3/1', 'xe-2/1/2', 'xe-2/1/3', 'xe-2/2/2', 'xe-2/2/3', 'xe-2/3/3', 'xe-3/0/2', 'xe-3/1/2', 'xe-3/2/1', 'xe-3/2/2']}
avail_port_dict_z = {'icn54-54-vc-svc-r311':["xe-0/0/0:0","xe-0/0/0:1","xe-0/0/0:2","xe-0/0/0:3","xe-0/0/8:0","xe-0/0/8:1","xe-0/0/8:2","xe-0/0/8:3","xe-0/0/16:0","xe-0/0/16:1","xe-0/0/16:2","xe-0/0/16:3","xe-0/0/24:0","xe-0/0/24:1","xe-0/0/24:2","xe-0/0/24:3"], 'icn54-54-vc-svc-r312':["xe-0/0/0:0","xe-0/0/0:1","xe-0/0/0:2","xe-0/0/0:3","xe-0/0/8:0","xe-0/0/8:1","xe-0/0/8:2","xe-0/0/8:3","xe-0/0/16:0","xe-0/0/16:1","xe-0/0/16:2","xe-0/0/16:3","xe-0/0/24:0","xe-0/0/24:1","xe-0/0/24:2","xe-0/0/24:3"]}


parser = argparse.ArgumentParser(description="Create cutsheet file for scaling/migration")
parser.add_argument("-ahosts", "--hostnames_a",
                    help="list of hostnames in A side, <device 1>,<device 2>,...,<device n>")
parser.add_argument("-zhosts", "--hostnames_z",
                    help="list of hostnames in Z side, <device 1>,<device 2>,...,<device n>")
parser.add_argument("-ainum", "--int_count_a", type=int, choices = [16,32],
                    help="Number of interfaces needed per device in A side")
parser.add_argument("-zinum", "--int_count_z", type=int, default = None, choices = [16,32],
                    help="Number of interfaces needed per device in Z side, default value is the same as A side")
    
args = parser.parse_args()

print(args)
    
if not args.int_count_z:
    args.int_count_z = args.int_count_a

args.hostnames_a = args.hostnames_a.split(',')
args.hostnames_z = args.hostnames_z.split(',')



wb_cutsheets = Workbook()
ws_cutsheet = wb_cutsheets.active
    
#avail_port_dict_z = svc_available_ports(args.int_count_z,args.hostnames_z)
z_int_idx = xlsx_styles.headers.index('z_interface')
z_host_idx = xlsx_styles.headers.index('z_hostname')
    
initialize_xlsx(ws_cutsheet, args.int_count_a, args.hostnames_a)

datafill_host_int(hostnames = args.hostnames_a, int_count = args.int_count_a, avail_port_dict = avail_port_dict_a, ws_cutsheet = ws_cutsheet)
datafill_host_int(host_col_idx = z_host_idx, int_col_idx = z_int_idx, hostnames = args.hostnames_z, int_count = args.int_count_z, avail_port_dict = avail_port_dict_z, ws_cutsheet = ws_cutsheet)
    
date_str = datetime.now().strftime("%m_%d_%y")
spreadsheet_filename = args.hostnames_a[0][:-1] + '_edg_vegemite_' + date_str + '.xlsx' 
print(spreadsheet_filename)
wb_cutsheets.save(spreadsheet_filename)
