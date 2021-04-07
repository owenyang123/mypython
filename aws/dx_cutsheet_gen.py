#!/apollo/sbin/envroot "$ENVROOT/python3.6/bin/python3.6"

import os
import argparse
import logging
import getpass
import re
import csv
from datetime import date
from collections import OrderedDict
import openpyxl
from openpyxl.styles.borders import Border, Side
from nre_tools_nsm_client.nsm_client import NSMClient, NSMError
from dxd_tools_dev.modules.NetworkDevice import VcEdg, VcBar, BrAgg, BrKct, VcCor, VcCar
from dxd_tools_dev.portdata import border, portfunction


def parse_args():
    """Parse args input by user and help display"""
    parser = argparse.ArgumentParser(description='''Create cutsheets for DX new edge group deployment.\n
    wiki:https://w.amazon.com/bin/view/DXDEPLOY_Automation_Others/dx_cutsheet_gen.py
    Examples:
    /apollo/env/DXDeploymentTools/bin/dx_cutsheet_gen.py -az syd5 -nz 51 -eg 2 -type new-edge-group
    /apollo/env/DXDeploymentTools/bin/dx_cutsheet_gen.py -az nrt55 -nz 55 -eg 5,6 -type hambone-v2 -er dori-s -links 8  
    ''', add_help=True, formatter_class=argparse.RawTextHelpFormatter)
    parser.add_argument('-type', '--cutsheet-type', required=True, dest='type',choices=['hambone-v2', 'new-edge-group', 'phoenix-v2-cor'], help='Type of deployment')
    parser.add_argument('-az', '--az',  help='e.g. nrt55,syd7')
    parser.add_argument('-nz', '--nz',  help='e.g. for syd5-51, nz is 51')
    parser.add_argument('-pop', '--pop', help='For POP deployments, e.g. arn53,syd1')
    parser.add_argument('-eg', '--edge-group', dest='edge_group',
                        help='Edge Group number. Either a number or a csv of numbers.'
                             'Maximum 4 groups supported. e.g. 3\n e.g. 4,5,6')
    parser.add_argument('-agg-links', '--br-agg-no-links', dest='br_agg_links', type=int, default=4,
                        help='Number of 10G links between 1xVC-BAR and 1xBR-AGG. For example, if this variable is\n'
                              'set to 4 with 2xVC-BARs and 4xBR-AGGs, total inter layer capacity is 4x10Gx2x4 = 320Gb')
    parser.add_argument('-car-links', '--vc-car-no-links', dest='vc_car_links', type=int, default=1, choices= [1,2,3],
                        help='Number of 100G links between 1xVC-CAR and 1xVC-COR. Default is 1. For example, if this variable is\n'
                              'set to 3, total inter layer capacity is 3x100Gx2x4 = 2.4Tb')
    parser.add_argument('-er', '--edge-rack', dest='edge_rack', choices=['vegemite', 'dori-s'], default='vegemite',
                        help='Edge Rack type, Default=vegemite')
    parser.add_argument('-4-bar', '--4-wide-bar', dest='vc_bar_count', action='store_true',
                        help='Specify this if cutsheet requrie 4 VC-BARs. Does not take arguments and if not specified,\n'
                             'default is 2')
    parser.add_argument('-bar', '--vc-bar-names', dest='vc_bar_names',
                        help='Only needed when there are more VC-BARs in NSM than required for the cutsheet.\n'
                             'Input VC-BAR hostnames or numbers, e.g. iad12-vc-bar-r5,iad12-vc-bar-r6. e.g. r5,r6')
    return parser.parse_args()

def setup_logging():
    """
    Setup logging messages to include time and different colors for logging levels
    ref: https://code.amazon.com/packages/ISDToolsScripts/blobs/292816ceeda05d6ee304fa37483e9da33d68f3f9/--/bin/gbwrap.py#L55
    """
    logging.basicConfig(format='%(asctime)s - %(levelname)s - %(message)s', level=logging.INFO)
    logging.addLevelName(logging.INFO, "\033[36m%s\033[1;0m" % logging.getLevelName(logging.INFO))
    logging.addLevelName(logging.WARNING, "\033[33m%s\033[1;0m" % logging.getLevelName(logging.WARNING))
    logging.addLevelName(logging.ERROR, "\033[31m%s\033[1;0m" % logging.getLevelName(logging.ERROR))


def validate_args(args):
    """validates user input and parse some args"""
    logging.info('Validating Input')
    if args.vc_bar_count:
        args.vc_bar_count = 4
    else:
        args.vc_bar_count = 2


def sort_devices(devices):
    """
    sorts devices based on hostname
    :param devices: list of devices
    :return sorted list of devices
    """
    return sorted(devices, key=lambda x: str(x))

def vc_bar_count_check(devices, args):
    """
    Count number of VC-BARs, return error if none or more than expected VC-BARs exist and --vc-bar-names isn't specified.
    Still work in progress
    :param devices: a list of device objects, args:
    """
    if len(devices) == 0:
        logging.error('No VC-BARs could be found in AZ {}'.format(args.az))
        exit()
    if len(devices) > args.vc_bar_count:
        logging.error('Existing VC-BARs AZ {}:'.format(args.az))
        logging.error('More than 2 VC-BARs exist AZ {}. Please re-run script with --vc-bar argument'.format(args.az))
        exit()

def decrease_speed(speed):
    """
    Sets lower speed for assigning BR-KCT interfaces
    :param speed:
    :return:
    """
    if speed == '40000':
        return '10000'
    elif speed == '10000':
        return None


def count_intf(devices, free_intf_speed):
    """
    Used for BR-KCTs to check enough interfaces exist to support required bandwidth
    :param devices:
    :param free_intf_speed:
    :return:
    """
    speed_in_gb = str(int(int(free_intf_speed)/1000)) + 'G'
    required_bandwidth = 160000
    required_links = int(required_bandwidth/int(free_intf_speed))
    new_free_intf_speed = 'same'
    logging.info('Validating sufficient number of free {} interfaces'.format(speed_in_gb))
    for device in devices:
        if device.free_vpc_intfs.get(free_intf_speed):
            if len(device.free_vpc_intfs[free_intf_speed]) > required_links:
                logging.info('{} has sufficient free {} interfaces. Required = {}, Free = {}'.format(device.name, speed_in_gb, required_links,len(device.free_vpc_intfs[free_intf_speed])))
            else:
                logging.warning('{} does not have sufficient free {} interfaces. Required = {}, Free = {}'.format(device.name, speed_in_gb, required_links,len(device.free_vpc_intfs[free_intf_speed])))
                new_free_intf_speed = decrease_speed(free_intf_speed)
        else:
            logging.warning(
                '{} does not have any free {} interfaces. Required = {}, Free = 0'.format(device.name, speed_in_gb,
                                                                                              required_links))
            new_free_intf_speed = decrease_speed(free_intf_speed)
    if new_free_intf_speed == 'same':
        return free_intf_speed
    else:
        return new_free_intf_speed


def check_intf_count(devices):
    """
    Checks if BR-KCTs have sufficient 40G interfaces(4 per KCT), else checks for sufficient 10G interfaces(61 per KCT)
    :param devices:
    :return:
    """
    free_intf_speed = '40000'
    free_intf_speed = count_intf(devices, free_intf_speed)
    if free_intf_speed == '10000':
        free_intf_speed = count_intf(devices, free_intf_speed)
    if free_intf_speed is None:
        logging.error('Not enough BR-KCT interfaces for deployment')
        exit()
    return free_intf_speed


def create_cell_border(sheet, row, first_column, last_column):
    """Create borders for given cells"""
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for column_no in range(first_column, last_column+1):
        sheet.cell(row=row, column=column_no).border = thin_border


# Code optimization: can possibly combine create_cover_sheet methods into one method


def create_cover_sheet_hambone(cutsheet, args):
    """
    Create cover sheet for hambone_v2 and new vc-edge groups
    :param cutsheet: cutsheet excel object
    :param args:
    """
    logging.info('Creating Cover Sheet')
    cover_sheet = cutsheet.get_sheet_by_name('Cover Sheet')
    cover_sheet['B1'].value = args.az
    cover_sheet['B2'].value = args.nz
    cover_sheet['B3'].value = getpass.getuser()
    cover_sheet['B4'].value = date.today().strftime("%d-%b-%Y")


def create_cover_sheet_phoenix(cutsheet, args, region):
    """
        Create cover sheet for phoenix_v2
        :param cutsheet: cutsheet excel object
        :param args:
    """
    logging.info('Creating Cover Sheet')
    cover_sheet = cutsheet.get_sheet_by_name('Cover Sheet')
    cover_sheet['B1'].value = args.pop
    cover_sheet['B2'].value = region
    cover_sheet['B3'].value = getpass.getuser()
    cover_sheet['B4'].value = date.today().strftime("%d-%b-%Y")

# code optimization: can potentially combine query_br_agg() and query_vc_bar() in one method


def query_br_kct(args):
    """
    Query NSM for BR-KCTs and port availability data
    :param args:
    :return a list of BR-KCT objects
    """
    br_kcts = []
    try:
        logging.info('Querying NSM for BR-KCTs in {}'.format(args.pop.upper()))
        nsm_client = NSMClient()
        br_kct_nsm = nsm_client.search_for_devices(pattern='name:/{}-br-kct.*/'.format(args.pop))
    except NSMError:
        logging.error(NSMError)
        exit()
        
    for i, kct in enumerate(br_kct_nsm):
        if kct.lifecycle_status != 'OPERATIONAL' and kct.lifecycle_status != 'TURNED_UP':
            logging.warning('{} is in {} NSM state. Ignoring..'.format(kct.name, kct.lifecycle_status))
        else:
            br_kcts.append(BrKct(kct.name))
            logging.info('Searching port availability data for free interfaces on {}'.format(br_kcts[i].name))
            br_kcts[i].get_free_intf(border.get_device_available_port(kct.name))
    return br_kcts, br_kct_nsm[0].region


def query_br_agg(args):
    """
    Query NSM for BR-AGGs and ISDtools port availability data
    :param args:
    :return a list of BR-AGG objects
    """
    br_aggs = []
    try:
        logging.info('Querying NSM for BR-AGGs in {}'.format(args.az.upper()))
        nsm_client = NSMClient()
        br_agg_nsm = nsm_client.search_for_devices(pattern='name:/{}-br-agg-r.*/'.format(args.az))
    except NSMError:
        logging.error(NSMError)
        exit()
    for i, br_agg in enumerate(br_agg_nsm):
        if br_agg.lifecycle_status != 'OPERATIONAL' and br_agg.lifecycle_status != 'TURNED_UP':
            logging.warning('{} is in {} NSM state. Ignoring..'.format(br_agg.name, br_agg.lifecycle_status))
        else:
            br_aggs.append(BrAgg(br_agg.name, br_agg.model))
            logging.info('Searching port availability data for free interfaces on {}'.format(br_aggs[i].name))
            br_aggs[i].get_free_intf(border.get_device_available_port(br_aggs[i].name)['10000'])
    return br_aggs


def query_vc_bar(args):
    """
    Query NSM for VC-BARs and port availability data
    :param args:
    :return a list of VC-BARs objects
    """
    vc_bars = []
    try:
        logging.info('Querying NSM for VC-BARS in {}'.format(args.az.upper()))
        nsm_client = NSMClient()
        if args.vc_bar_names:  # need code to acommodate 4 vc-bars
            vc_bar_nsm = nsm_client.search_for_devices(
                pattern='name:/{}|{}/'.format(args.vc_bar_names.split(',')[0], args.vc_bar_names.split(',')[1]))
        else:
            vc_bar_nsm = nsm_client.search_for_devices(pattern='name:/{}-vc-bar-r.*/'.format(args.az))
    except NSMError:
        logging.error(NSMError)
        exit()
    vc_bar_nsm = sort_devices(vc_bar_nsm)
    vc_bar_count_check(vc_bar_nsm, args)
    for i, vc_bar in enumerate(vc_bar_nsm):
        vc_bars.append(VcBar(vc_bar.name))
        portfunction.get_device_available_port(vc_bar.name)
        logging.info('Searching port availability data for free interfaces on {}'.format(vc_bar.name))
        vc_bars[i].get_free_intf(portfunction.get_device_available_port(vc_bar.name)['10000'])

    for vc_bar in vc_bars:  # find common interfaces between VC-BARs to maintain symmetry of assigned interfaces
        vc_bars[0].find_common_intf(vc_bar)
        vc_bar.find_common_intf(vc_bars[0])
    return vc_bars


def create_vc_bar_br_agg_sheet(cutsheet, br_aggs, vc_bars, args):
    """
    Create sheet for VC-BAR to BR-AGG connections
    :param cutsheet: cutsheet excel object
    :param br_aggs: list of BR-AGG objects
    :param vc_bars: list of VC-BAR objects
    :param args:
    """
    logging.info('Creating vc-bar<>br-agg sheet')
    for vc_bar in vc_bars:
        vc_bar.select_free_intf('border', 'new')
    br_agg_to_vc_bar = cutsheet.get_sheet_by_name('vc-bar<>br-agg')
    row_index = 3
    for vc_bar in vc_bars:
        for br_agg_index, br_agg in enumerate(br_aggs):
            for link_index in range(0, args.br_agg_links):
                vc_bar.assign_free_intf = ['border', br_agg_index]
                br_agg_intf = br_aggs[br_agg_index].assign_free_intf
                br_agg_to_vc_bar.cell(row=row_index, column=35).value = br_agg_intf.name
                br_agg_to_vc_bar.cell(row=row_index, column=36).value = br_agg.name
                br_agg_to_vc_bar.cell(row=row_index, column=2).value = vc_bar.assign_free_intf.name
                br_agg_to_vc_bar.cell(row=row_index, column=1).value = vc_bar.name
                create_cell_border(br_agg_to_vc_bar, row_index, 1, 38)
                row_index += 1
                if link_index + 1 == args.br_agg_links:  # ensure that 2 VC-BARs don't get assigned same FPC on BR-AGG
                    if br_agg_intf.fpc == br_agg.free_intf_by_fpc[0][0].fpc:
                        br_agg.free_intf_by_fpc.pop(0)  # better to pop it to end of list ;)
        create_cell_border(br_agg_to_vc_bar, row_index, 1, 38)
        row_index += 1


def create_vc_bar_vc_edg_sheet(cutsheet, vc_bars, args):
    """
    Create sheet for VC-BAR to VC-EDG connections
    :param cutsheet: cutsheet excel object
    :param vc_bars: list of VC-BAr objects
    :param args:
    """
    logging.info('Creating vc-bar<>vc-edg sheet')
    vc_bar_to_vc_edg = cutsheet.get_sheet_by_name('vc-bar<>vc-edg')
    no_links = int(32 / len(vc_bars))
    row_index = 3
    for vc_bar_index, vc_bar in enumerate(vc_bars):
        if args.type == 'hambone-v2':
            vc_bar.select_free_intf('vpc', 'new')
        if args.type == 'new-edge-group':
            vc_bar.select_free_intf('vpc', 'operational')
        for eg_index, edge_group in enumerate(args.edge_group):
            vc_edgs = [VcEdg('{}-vc-edg-r3{}1'.format(args.az, args.edge_group[eg_index] - 1)),
                       VcEdg('{}-vc-edg-r3{}2'.format(args.az, args.edge_group[eg_index] - 1))]
            vc_edgs[0].select_free_vpc_intfs()
            vc_edgs[1].select_free_vpc_intfs()
            for vc_edg_index, vc_edg in enumerate(vc_edgs):
                if vc_bar_index % 2 == vc_edg_index:
                    vc_edg.assign_free_vpc_intf = 0
                else:
                    vc_edg.assign_free_vpc_intf = 1
                vc_edg.select_free_vpc_intfs()
                for link_index in range(0, no_links):
                    vc_bar.assign_free_intf = ['vpc', vc_edg_index]
                    vc_bar_to_vc_edg.cell(row=row_index, column=1).value = vc_bar.name
                    vc_bar_to_vc_edg.cell(row=row_index, column=2).value = vc_bar.assign_free_intf.name
                    vc_bar_to_vc_edg.cell(row=row_index, column=35).value = vc_edg.assign_free_vpc_intf.name
                    vc_bar_to_vc_edg.cell(row=row_index, column=36).value = vc_edg.name
                    create_cell_border(vc_bar_to_vc_edg, row_index, 1, 38)
                    row_index += 1
                create_cell_border(vc_bar_to_vc_edg, row_index, 1, 38)
                row_index += 1


def create_vc_edg_svc_sheet(cutsheet, args):
    """
    Create sheet for VC-EDG to SVC connections
    :param cutsheet: cutsheet excel object
    :param args:
    """
    if args.edge_rack.lower() == 'vegemite':
       cutsheet.remove_sheet(cutsheet.get_sheet_by_name('vc-edg<>es-svc'))
       vc_edg_to_svc = cutsheet.get_sheet_by_name('vc-edg<>vc-svc')
    elif args.edge_rack.lower() == 'dori-s':
       cutsheet.remove_sheet(cutsheet.get_sheet_by_name('vc-edg<>vc-svc'))
       vc_edg_to_svc = cutsheet.get_sheet_by_name('vc-edg<>es-svc')
    for eg_count, edge_group in enumerate(args.edge_group):
        vc_edg_to_svc_copy = cutsheet.copy_worksheet(vc_edg_to_svc)
        vc_edg_to_svc_copy.title = 'vc-edg<>svc (EG{})'.format(edge_group)
        vc_edgs = [VcEdg('{}-vc-edg-r3{}1'.format(args.az, args.edge_group[eg_count] - 1)),
                   VcEdg('{}-vc-edg-r3{}2'.format(args.az, args.edge_group[eg_count] - 1))]
        for link_index in range(0, 16):
            if args.edge_rack.lower() == 'vegemite':  # code candidate for trimming
                blackfoot_hostname_1 = '{}-{}-vc-svc-r3{}1'.format(args.az, args.nz, args.edge_group[eg_count] - 1)
                blackfoot_hostname_2 = '{}-{}-vc-svc-r3{}2'.format(args.az, args.nz, args.edge_group[eg_count] - 1)
                vc_edg_to_svc_copy.cell(row=link_index + 3, column=1).value = vc_edgs[0].name
                vc_edg_to_svc_copy.cell(row=link_index + 3, column=36).value = blackfoot_hostname_1
                vc_edg_to_svc_copy.cell(row=link_index + 20, column=1).value = vc_edgs[1].name
                vc_edg_to_svc_copy.cell(row=link_index + 20, column=36).value = blackfoot_hostname_2
            elif args.edge_rack.lower() == 'dori-s':
                if link_index < 9:
                    blackfoot_hostname = '{}-{}-es-svc-r3{}10{}'.format(args.az, args.nz, args.edge_group[eg_count]-1, link_index+1)
                else:
                    blackfoot_hostname = '{}-{}-es-svc-r3{}1{}'.format(args.az, args.nz, args.edge_group[eg_count] - 1, link_index+1)
                vc_edg_to_svc_copy.cell(row=link_index+3, column=1).value = vc_edgs[0].name
                vc_edg_to_svc_copy.cell(row=link_index+3, column=36).value = blackfoot_hostname
                vc_edg_to_svc_copy.cell(row=link_index+20, column=1).value = vc_edgs[1].name
                vc_edg_to_svc_copy.cell(row=link_index+20, column=36).value = blackfoot_hostname
    cutsheet.remove_sheet(vc_edg_to_svc)


def create_vc_cir_vc_bar_sheet(cutsheet, vc_bars):
    """
    Create sheet for VC-CIR to VC-BAR connections
    :param cutsheet: cutsheet excel object
    :param vc_bars:  list of VC-BAr objects
    """
    logging.info('Creating vc-cir<>vc-bar sheet')
    vc_cir_to_vc_bar = cutsheet.get_sheet_by_name('vc-cir<>vc-bar')
    no_links = int(16 / len(vc_bars))
    row_index = 3
    for x in range(0, 2):
        for vc_bar_index, vc_bar in enumerate(vc_bars):
            vc_bar.select_free_intf('cir', 'new')  # Assumption is vc-cirs are only deployed with new hambone-v2
            for i in range(0, no_links):
                if vc_bar_index % 2 == x:
                    vc_bar.assign_free_intf = ['cir', 0]
                else:
                    vc_bar.assign_free_intf = ['cir', 1]
                vc_cir_to_vc_bar.cell(row=row_index, column=35).value = vc_bar.assign_free_intf.name
                vc_cir_to_vc_bar.cell(row=row_index, column=36).value = vc_bar.name
                row_index += 1
        row_index += 1


def create_vc_cor_br_kct_sheet(cutsheet, vc_cors, br_kcts, free_intf_speed):
    """
     Create sheet for VC-COR to BR-KCT connections
    :param cutsheet:
    :param args:
    :param br_kcts:
    :param free_intf_speed:
    :return:
    """
    logging.info('Creating vc-cor<>br-kct sheet')
    vc_cor_to_br_kct = cutsheet.get_sheet_by_name('vc-cor<>br-kct')
    row_index = 2
    for vc_cor in vc_cors:
        vc_cor.border_intf = free_intf_speed
        vc_cor.select_free_intf('border', 'new', None)
        for kct in br_kcts:
            for i in range(0, int(40000 / int(free_intf_speed))):  # iterate once for 40G and 4 times for 10G interfaces
                kct.assign_free_intf = free_intf_speed
                kct.get_speed(free_intf_speed)
                vc_cor.lag = kct.name
                kct.lag = vc_cor.name
                vc_cor_to_br_kct.cell(row=row_index, column=1).value = vc_cor.name
                vc_cor_to_br_kct.cell(row=row_index, column=2).value = vc_cor.lag
                vc_cor_to_br_kct.cell(row=row_index, column=3).value = vc_cor.assign_free_intf.name
                vc_cor_to_br_kct.cell(row=row_index, column=4).value = vc_cor.optic
                vc_cor_to_br_kct.cell(row=row_index, column=5).value = vc_cor.connector
                vc_cor_to_br_kct.cell(row=row_index, column=6).value = 'SMF'
                vc_cor_to_br_kct.cell(row=row_index, column=7).value = kct.connector
                vc_cor_to_br_kct.cell(row=row_index, column=8).value = kct.optic
                vc_cor_to_br_kct.cell(row=row_index, column=9).value = kct.assign_free_intf.name
                vc_cor_to_br_kct.cell(row=row_index, column=10).value = kct.lag
                vc_cor_to_br_kct.cell(row=row_index, column=11).value = kct.name
                create_cell_border(vc_cor_to_br_kct, row_index, 1, 11)
                row_index += 1
        create_cell_border(vc_cor_to_br_kct, row_index, 1, 11)
        row_index += 1


def create_vc_cor_vc_car(cutsheet, args, vc_cors):
    """
    Create sheet for VC-COR to VC-CAR connections
    :param cutsheet: 
    :param args: 
    :return: 
    """
    logging.info('Creating vc-cor<>vc-car sheet')
    vc_cor_to_vc_car = cutsheet.get_sheet_by_name('vc-cor<>vc-car')
    vc_cars = []
    for i in range(0, 2):
        vc_cars.append(VcCar('{}-vc-car-p2-v3-r{}'.format(args.pop, i+1)))
    row_index = 2
    for vc_cor_index, vc_cor in enumerate(vc_cors):
        for vc_car_index, vc_car in enumerate(vc_cars):
            vc_car.select_free_intfs()
            vc_car.assign_free_intf = vc_cor_index
            vc_car.lag = vc_cor.name
            vc_cor.lag = vc_car.name
            vc_cor.select_free_intf('car', 'new', vc_car_index)
            for i in range(0, args.vc_car_links):
                vc_cor_to_vc_car.cell(row=row_index, column=1).value = vc_cor.name
                vc_cor_to_vc_car.cell(row=row_index, column=2).value = vc_cor.lag
                vc_cor_to_vc_car.cell(row=row_index, column=3).value = vc_cor.assign_free_intf.name
                vc_cor_to_vc_car.cell(row=row_index, column=4).value = 'QSFP28 100G CWDM'
                vc_cor_to_vc_car.cell(row=row_index, column=5).value = 'LC'
                vc_cor_to_vc_car.cell(row=row_index, column=6).value = 'SMF'
                vc_cor_to_vc_car.cell(row=row_index, column=7).value = 'LC'
                vc_cor_to_vc_car.cell(row=row_index, column=8).value = 'QSFP28 100G CWDM'
                vc_cor_to_vc_car.cell(row=row_index, column=9).value = vc_car.assign_free_intf.name
                vc_cor_to_vc_car.cell(row=row_index, column=10).value = vc_car.lag
                vc_cor_to_vc_car.cell(row=row_index, column=11).value = vc_car.name
                create_cell_border(vc_cor_to_vc_car, row_index, 1, 11)
                row_index += 1
        create_cell_border(vc_cor_to_vc_car, row_index, 1, 11)
        row_index += 1


def create_gbwrap_csv(cutsheet, output_directory, args):
    """
    Creates csv that can be used by gbwrap for port reservations
    LAG numbering is based on observed previous deployments, also:
    wiki: https://w.amazon.com/bin/view/Networking/IS/Reference/Standards/Juniper_AE_Numbering/#Data_Center
    :param cutsheet:
    :param output_directory:
    :param args:
    :return:
    """
    gbrwap_csv = os.path.join(output_directory, '{}_br_agg_gbwrap_portres.csv'.format(args.az))
    logging.info('Creating {}'.format(gbrwap_csv))
    logging.warning('BR-AGG LAG numbers in {} are assigned statically. Please verify on devices'.format(gbrwap_csv))
    sheet = cutsheet.get_sheet_by_name('vc-bar<>br-agg')
    empty_count = 0
    with open(gbrwap_csv, mode='w') as port_res:
        writer = csv.writer(port_res, delimiter=',', quoting=csv.QUOTE_NONE)
        writer.writerow(['a_hostname', 'a_interface', 'a_lag', 'z_lag', 'z_interface', 'z_hostname'])
        for row in range(3, sheet.max_row):
            a_hostname = sheet.cell(row=row, column=1).value
            a_intf = sheet.cell(row=row, column=2).value
            z_hostname = sheet.cell(row=row, column=36).value
            z_intf = sheet.cell(row=row, column=35).value
            if a_hostname is not None:
                a_lag = 'ae' + str(int(re.match(r'\w+\d+-br-agg-r(\d+)', z_hostname).group(1)) + 19)
                z_lag = 'ae' + str(int(re.match(r'\w+\d+-vc-bar-r(\d+)', a_hostname).group(1)) + 22)
                writer.writerow([a_hostname, a_intf, a_lag, z_lag, z_intf, z_hostname])
            else:
                empty_count += 1
                if empty_count >= 4:
                    break


def save_cutsheet(cutsheet, output_directory, args):
    """
    Given a cutsheet and deployment type, saves cutsheet to specified directory
    :param cutsheet: 
    :param output_directory: 
    :param args: 
    :return: 
    """
    if not os.path.exists(output_directory):
        logging.info('Directory {} doesn\'t exist. Creating..'.format(output_directory))
        os.mkdir(output_directory)
    if args.type == 'hambone-v2':
        cutsheet_name = os.path.join(output_directory, '{}_Hambone_v2_cutsheet.xlsx'.format(args.az))
    elif args.type == 'new-edge-group':
        edge_groups = ''
        for group in args.edge_group:
            edge_groups = edge_groups + '{}_'.format(group)
        cutsheet_name = os.path.join(output_directory, '{}_Edge_Group_{}cutsheet.xlsx'.format(args.az, edge_groups))
    elif args.type == 'phoenix-v2-cor':
        cutsheet_name = os.path.join(output_directory, '{}_phoenix_v2_cutsheet.xlsx'.format(args.pop))
    logging.info('Creating {}'.format(cutsheet_name))
    cutsheet.save('{}'.format(cutsheet_name))


def create_cutsheet(args):
    """
    Create cutsheet
    :param args:
    """
    if args.type == 'hambone-v2' or args.type == 'new-edge-group':
        template = 'hambone_v2_cutsheet_template.xlsx'
        args.edge_group = list(map(int, args.edge_group.split(',')))
    elif args.type == 'phoenix-v2-cor':
        template = 'phoenix_v2_full_cutsheet_template.xlsx'
    # template_path = '/home/abaseer/DXDeploymentTools/src/DXDeploymentTools/configuration/cutsheet_templates/'
    template_path = '/apollo/env/DXDeploymentTools/cutsheet_templates/'
    cutsheet = openpyxl.load_workbook(os.path.join(template_path, template))
    output_directory = os.path.join(os.environ['HOME'], 'dx_cutsheet_gen_out')
    if args.type == 'hambone-v2':
        vc_bars = []
        for vc_bar_count in range(0, args.vc_bar_count):
            vc_bars.append(VcBar('{}-vc-bar-r{}'.format(args.az, vc_bar_count + 1)))
        create_cover_sheet_hambone(cutsheet, args)
        create_vc_bar_br_agg_sheet(cutsheet, query_br_agg(args), vc_bars, args)
        create_vc_bar_vc_edg_sheet(cutsheet, vc_bars, args)
        create_vc_edg_svc_sheet(cutsheet, args)
        create_vc_cir_vc_bar_sheet(cutsheet, vc_bars)
        logging.info('Creating remaining sheets')
        cutsheet.active = 0
        save_cutsheet(cutsheet, output_directory, args)
        create_gbwrap_csv(cutsheet, output_directory, args)
    elif args.type == 'new-edge-group':
        create_cover_sheet_hambone(cutsheet, args)
        cutsheet.remove_sheet(cutsheet.get_sheet_by_name('vc-bar<>br-agg'))
        cutsheet.remove_sheet(cutsheet.get_sheet_by_name('vc-cir<>vc-bar'))
        cutsheet.remove_sheet(cutsheet.get_sheet_by_name('vc-cir<>vc-xlc'))
        cutsheet.remove_sheet(cutsheet.get_sheet_by_name('xlc Loopback'))
        create_vc_bar_vc_edg_sheet(cutsheet, query_vc_bar(args), args)
        create_vc_edg_svc_sheet(cutsheet, args)
        logging.info('Creating remaining sheets')
        cutsheet.active = 0
        save_cutsheet(cutsheet, output_directory, args)
    elif args.type == 'phoenix-v2-cor':
        br_kcts, region = query_br_kct(args)
        free_intf_speed = check_intf_count(br_kcts)
        vc_cors = []
        for i in range(0, 4):
            vc_cors.append(VcCor('{}-vc-cor-b1-r{}'.format(args.pop, i + 1)))
        create_cover_sheet_phoenix(cutsheet, args, region)
        create_vc_cor_br_kct_sheet(cutsheet, vc_cors, br_kcts, free_intf_speed)
        create_vc_cor_vc_car(cutsheet, args, vc_cors)
        logging.info('Creating remaining sheets')
        cutsheet.active = 0
        save_cutsheet(cutsheet, output_directory, args)

def main():
    args = parse_args()
    setup_logging()
    validate_args(args)
    create_cutsheet(args)


if __name__ == "__main__":
        main()

