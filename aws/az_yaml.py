#!/apollo/sbin/envroot "$ENVROOT/python3.6/bin/python3.6"

import logging
from datetime import datetime
import re
import os
import argparse
from collections import OrderedDict, Mapping
import openpyxl
import yaml
from random import randint, seed
import jinja2
from dxd_tools_dev.modules.utils import setup_logging, validate_file_exists


class InetCidr(object):
    """
    CIDR table class as per standards from wikis
    wiki: https://w.amazon.com/bin/view/Interconnect/HAMBonev2#Addressing
    https://w.amazon.com/bin/view/Networking/Piyushsi/Customer-DX_Deployment_Run_Book/#H3.1ReserveIPaddress
    """

    def __init__(self):
        self.cidr_table = [
        ['edg-r311', 'bar-r1', '169.254.211.10/31'],
        ['edg-r311', 'bar-r2', '169.254.211.12/31'],
        ['edg-r312', 'bar-r1', '169.254.211.14/31'],
        ['edg-r312', 'bar-r2', '169.254.211.16/31'],
        ['edg-r321', 'bar-r1', '169.254.211.18/31'],
        ['edg-r321', 'bar-r2', '169.254.211.20/31'],
        ['edg-r322', 'bar-r1', '169.254.211.22/31'],
        ['edg-r322', 'bar-r2', '169.254.211.24/31'],
        ['edg-r331', 'bar-r1', '169.254.211.26/31'],
        ['edg-r331', 'bar-r2', '169.254.211.28/31'],
        ['edg-r332', 'bar-r1', '169.254.211.30/31'],
        ['edg-r332', 'bar-r2', '169.254.211.32/31'],
        ['edg-r341', 'bar-r1', '169.254.211.34/31'],
        ['edg-r341', 'bar-r2', '169.254.211.36/31'],
        ['edg-r342', 'bar-r1', '169.254.211.38/31'],
        ['edg-r342', 'bar-r2', '169.254.211.40/31'],
        ['cir-r1', 'bar-r1', '169.254.211.42/31'],
        ['cir-r1', 'bar-r2', '169.254.211.44/31'],
        ['cir-r2', 'bar-r1', '169.254.211.46/31'],
        ['cir-r2', 'bar-r2', '169.254.211.48/31'],
        ['cir-r1', 'bar-r1', '169.254.211.50/31'],
        ['cir-r1', 'bar-r2', '169.254.211.52/31'],
        ['cir-r2', 'bar-r1', '169.254.211.54/31'],
        ['cir-r2', 'bar-r2', '169.254.211.56/31'],
        ['edg-r311', 'bar-r3', '169.254.211.122/31'],
        ['edg-r311', 'bar-r4', '169.254.211.124/31'],
        ['edg-r312', 'bar-r3', '169.254.211.126/31'],
        ['edg-r312', 'bar-r4', '169.254.211.128/31'],
        ['edg-r321', 'bar-r3', '169.254.211.130/31'],
        ['edg-r321', 'bar-r4', '169.254.211.132/31'],
        ['edg-r322', 'bar-r3', '169.254.211.134/31'],
        ['edg-r322', 'bar-r4', '169.254.211.136/31'],
        ['edg-r331', 'bar-r3', '169.254.211.138/31'],
        ['edg-r331', 'bar-r4', '169.254.211.140/31'],
        ['edg-r332', 'bar-r3', '169.254.211.142/31'],
        ['edg-r332', 'bar-r4', '169.254.211.144/31'],
        ['edg-r341', 'bar-r3', '169.254.211.146/31'],
        ['edg-r341', 'bar-r4', '169.254.211.148/31'],
        ['edg-r342', 'bar-r3', '169.254.211.150/31'],
        ['edg-r342', 'bar-r4', '169.254.211.152/31'],
        ['edg-r351', 'bar-r1', '169.254.211.154/31'],
        ['edg-r351', 'bar-r2', '169.254.211.156/31'],
        ['edg-r351', 'bar-r3', '169.254.211.158/31'],
        ['edg-r351', 'bar-r4', '169.254.211.160/31'],
        ['edg-r352', 'bar-r1', '169.254.211.162/31'],
        ['edg-r352', 'bar-r2', '169.254.211.164/31'],
        ['edg-r352', 'bar-r3', '169.254.211.166/31'],
        ['edg-r352', 'bar-r4', '169.254.211.168/31'],
        ['edg-r361', 'bar-r1', '169.254.211.170/31'],
        ['edg-r361', 'bar-r2', '169.254.211.172/31'],
        ['edg-r361', 'bar-r3', '169.254.211.174/31'],
        ['edg-r361', 'bar-r4', '169.254.211.176/31'],
        ['edg-r362', 'bar-r1', '169.254.211.178/31'],
        ['edg-r362', 'bar-r2', '169.254.211.180/31'],
        ['edg-r362', 'bar-r3', '169.254.211.182/31'],
        ['edg-r362', 'bar-r4', '169.254.211.184/31'],
        ['edg-r371', 'bar-r1', '169.254.211.186/31'],
        ['edg-r371', 'bar-r2', '169.254.211.188/31'],
        ['edg-r371', 'bar-r3', '169.254.211.190/31'],
        ['edg-r371', 'bar-r4', '169.254.211.192/31'],
        ['edg-r372', 'bar-r1', '169.254.211.194/31'],
        ['edg-r372', 'bar-r2', '169.254.211.196/31'],
        ['edg-r372', 'bar-r3', '169.254.211.198/31'],
        ['edg-r372', 'bar-r4', '169.254.211.200/31'],
        ['cir-r1', 'xlc-r101', '1.1.1.1/32'],
        ['cir-r2', 'xlc-r201', '1.1.1.1/32'],
        ['edg-r', 'svc', '10.0.0.2/31'],
        ]

    def get_cidr(self, a_hostname, z_hostname):
        for link in self.cidr_table:
            if re.search(link[0], a_hostname) and re.search(link[1], z_hostname):
                return link[2]
            elif re.search(link[1], a_hostname) and re.search(link[0], z_hostname):
                return link[2]
            else:
               inet_cidr = None
        return inet_cidr


INET_CIDR = InetCidr()


def parse_args():
    """Parse args input by user and help display"""
    parser = argparse.ArgumentParser(description='''Create DX AZ Yaml file from Cutsheet.\n
    wiki:TBD
    Examples:
    /apollo/env/DXDeploymentTools/bin/az_yaml.py --cutsheet ~/dx_cutsheet_gen_out/nrt7_hambone-v2.xlsx 
    /apollo/env/DXDeploymentTools/bin/az_yaml.py --cutsheet ~/dx_cutsheet_gen_out/syd51_hambone-v2.xlsx --mini-yaml ~/DX_Yamls/syd51_mini.yaml  
    ''', add_help=True, formatter_class=argparse.RawTextHelpFormatter)
    parser.add_argument('-c', '--cutsheet', required=True, dest='cutsheet', help='Path to cutsheet, e.g. ~/dx_cutsheet_gen_out/nrt7_hambone-v2.xlsx')
    parser.add_argument('-mini', '--mini-yaml',  dest='mini', help='Path to mini.yaml if existing , e.g.~/DX_Yamls/syd51_mini.yaml')
    return parser.parse_args()


def validate_input(args):
    """
    Validate user input
    :param args:
    :return:
    """
    logging.info('Valdating input')
    if not os.path.exists(os.path.join(os.environ['HOME'], 'DXYamls')):
        logging.error('DXYamls repo doesn\'t exist. please clone repo first "git clone ssh://git.amazon.com/pkg/DXYamls"')
        exit()
    validate_file_exists(args.cutsheet)
    if args.mini:
        validate_file_exists(args.mini)


def get_model(layer):
    """
    Returns default model for a certain layer
    :param layer:
    :return model:
    """
    if layer == 'VC-BAR':
        return'QFX10002-72Q'
    elif layer == 'VC-EDG':
        return'MX480'
    elif layer =='VC-CIR':
        return'MX240'
    elif layer == 'VC-XLC':
        return'QFX5100-96S'


def random_serial():
    """
    Creates random serial number based on current time seed
    :return:
    """
    seed(datetime.now())
    return 'TEMP' + str(randint(100000, 999999))


def parse_cir_xlc_hostname(az, hostname):
    """
    Parses and reformats hostname for VC-CIRs and VC-XLCs. Reason for this is openpyxel is not able to read formulas
    from excel sheet as far as I know
    :param az:
    :param hostname:
    :return hostname: reformatted hostname string
    """
    if re.match(r'.*(-vc-xlc-r\d+)', hostname):  # parse xlcs and cirs to remove formulas
        return az + re.match(r'.*(-vc-xlc-r\d+)', str(hostname)).group(1)
    elif re.match(r'.*(-vc-cir-r\d+)', hostname):
        return az + re.match(r'.*(-vc-cir-r\d+)', str(hostname)).group(1)
    else:
        return hostname


def get_z_hostname(a_hostname, hostname_pair):
    """
    Extracts the z_hostname from hostname_pair list of lists
    :param a_hostname:
    :param hostname_pair: a list of a list of hostname pairs e.g. [[nrt7-vc-bar-r1, nrt7-br-agg-r1], ...]
    :return z_hostname:
    """
    hostname_pair.remove(a_hostname)  # remove a_hostname
    return hostname_pair[0]  # return z_hostname


def create_hostname_pair(cutsheet):
    """
    extracts a list of uniques hostnames and hostname pairs from cutsheet
    :param cutsheet:
    :return hostname_pair: a list of a list of hostname pairs e.g. [[nrt7-vc-bar-r1, nrt7-br-agg-r1], ...]:
    :return hostnames: list of unique hostnames
    """
    logging.info('Parsing hostname pairs')
    hostname_pairs = []
    hostnames = []
    global AZ
    global NZ
    cover_sheet = cutsheet.get_sheet_by_name('Cover Sheet')
    AZ = cover_sheet.cell(row=1, column=2).value
    NZ = cover_sheet.cell(row=2, column=2).value
    cutsheet.remove_sheet(cutsheet.get_sheet_by_name('Cover Sheet'))
    for sheet in cutsheet.worksheets:
        empty_count = 0
        for row in range(3, sheet.max_row+3):  # +3 as a safety buffer to avoid a bug that ignored last row
            if sheet.cell(row=row, column=1).value is not None:  # Skips empty rows
                a_hostname = sheet.cell(row=row, column=1).value
                z_hostname = sheet.cell(row=row, column=36).value
                a_hostname = parse_cir_xlc_hostname(AZ, a_hostname)
                z_hostname = parse_cir_xlc_hostname(AZ, z_hostname)
                hostname_pairs.append((a_hostname, z_hostname))
                hostnames.append(a_hostname)
                hostnames.append(z_hostname)
                empty_count = 0
            else:
                empty_count += 1
                if empty_count >= 4:
                    break
    hostname_pairs = list(set(hostname_pairs))  # Creates unique pairs
    hostname_pairs = sorted(list(map(list, hostname_pairs)))
    return hostname_pairs, sorted(list(set(hostnames)))


def count_hostname(main_dict, hostnames_unique, hostname_pairs):
    """
    Counts how many devices exist in a layer as well as how many devices are connected to each device
    :param main_dict:
    :param hostnames_unique:
    :param hostname_pairs:
    :return count_dict: dictionary with count of a_hostnames and number of devices connected to them
    """
    logging.info('Counting hostname pairs')
    count_dict = OrderedDict()
    for hostname in hostnames_unique:  # Create layer and count number of a_hostnames
        try:
            layer = re.search(r'-(vc-(bar|edg|xlc|cir))-', hostname).group(1).upper()
            if main_dict['Devices'].get(layer) is None:  # Create layer if it doesn't exist
                main_dict['Devices'][layer] = OrderedDict()
                count_dict[layer] = OrderedDict()
                count_dict[layer]['count'] = 1
                count_dict[layer]['a_hostnames'] = [hostname]
            else:
                count_dict[layer]['count'] += 1
                count_dict[layer]['a_hostnames'].append(hostname)
                count_dict[layer]['a_hostnames'].sort()
        except:
            pass

    for hostname_pair in hostname_pairs:
        for hostname in hostname_pair:  # Count number of z_hostnames
            if count_dict.get(hostname) is None:  # and layer is not None:  # Create layer if it doesn't exist
                count_dict[hostname] = OrderedDict()
                count_dict[hostname] = 1
            else:
                count_dict[hostname] += 1
    return count_dict


def create_a_hostname(main_dict, count_dict):
    """
    Creates the dictionary structure for A-device
    :param main_dict:
    :param count_dict:
    """
    # logging.info('Creating a_device layer')
    for layer in main_dict['Devices'].keys():  # Create a_hostname dictionary
        for a_index in range(1, count_dict[layer]['count']+1):
            logging.info('Creating A-device: {}'.format(count_dict[layer]['a_hostnames'][a_index-1]))
            main_dict['Devices'][layer][a_index] = OrderedDict()
            main_dict['Devices'][layer][a_index]['a_hostname'] = count_dict[layer]['a_hostnames'][a_index-1]
            main_dict['Devices'][layer][a_index]['connections'] = OrderedDict()
            main_dict['Devices'][layer][a_index]['model'] = get_model(layer)
            main_dict['Devices'][layer][a_index]['serial'] = ''
            if layer == 'VC-EDG':
                main_dict['Devices'][layer][a_index]['woodchipper'] = ['xe-0/3/7', 'xe-1/3/7']


def create_z_hostname(main_dict, count_dict, hostname_pairs):
    """
    Creates the dictionary structure for connections
    :param main_dict:
    :param count_dict:
    :param hostname_pairs:
    :return:
    """
    hostname_pairs_copy = []  # used for second iteration
    for i in range(0,2):  # iterate twice to get z_hostnames for other end XLC->CIR, CIR->BAR
        layers = list(main_dict['Devices'].keys())
        if i == 1:
            layers.reverse()
            hostname_pairs = hostname_pairs_copy
        for layer in main_dict['Devices'].keys():
            for a_index in main_dict['Devices'][layer].keys():
                a_hostname = main_dict['Devices'][layer][a_index]['a_hostname']
                z_hostname_list = []  # used to make sure XLCs aren't included twice
                for z_index in range(1, count_dict[a_hostname]+1):
                    for hostname_pair in hostname_pairs:
                        if a_hostname in hostname_pair:
                            z_hostname = get_z_hostname(a_hostname, hostname_pairs.pop(hostname_pairs.index(hostname_pair)))
                            if a_hostname != z_hostname and z_hostname not in z_hostname_list:  # remove XLC<>XLC and redundant z_hostnames
                                logging.info('Creating connection: {} <> {}.'.format(a_hostname, z_hostname))
                                z_hostname_list.append(z_hostname)
                                hostname_pairs_copy.append([a_hostname, z_hostname])
                                main_dict['Devices'][layer][a_index]['connections'][z_index] = OrderedDict()
                                main_dict['Devices'][layer][a_index]['connections'][z_index]['z_hostname'] = z_hostname
                                main_dict['Devices'][layer][a_index]['connections'][z_index]['a_interfaces'] = []
                                main_dict['Devices'][layer][a_index]['connections'][z_index]['z_interfaces'] = []
                                main_dict['Devices'][layer][a_index]['connections'][z_index]['inet_cidr'] = INET_CIDR.get_cidr(a_hostname, z_hostname)
                                if re.match(r'\w+\d+-br-agg-r(\d+)', z_hostname):
                                    main_dict['Devices'][layer][a_index]['connections'][z_index]['z_lag'] = \
                                    'ae' + str(int(re.match(r'\w+\d+-vc-bar-r(\d+)', a_hostname).group(1)) + 22)
                                    main_dict['Devices'][layer][a_index]['connections'][z_index]['a_lag'] = \
                                        'ae' + str(int(re.match(r'\w+\d+-br-agg-r(\d+)', z_hostname).group(1)) + 20)
                                break


def create_intf(cutsheet, main_dict):
    """
    Fills info for a_interfaces and z_interfaces
    :param cutsheet:
    :param main_dict:
    :return:
    """
    logging.info('Parsing interfaces')
    for sheet in cutsheet.worksheets:  # this iteration for populating interfaces
        empty_count = 0
        for row in range(3, sheet.max_row+3):
            if sheet.cell(row=row, column=1).value is not None:  # Skips empty rows
                a_hostname = sheet.cell(row=row, column=1).value
                z_hostname = sheet.cell(row=row, column=36).value
                a_intf = sheet.cell(row=row, column=2).value
                z_intf = sheet.cell(row=row, column=35).value
                a_hostname = parse_cir_xlc_hostname(AZ, a_hostname)
                z_hostname = parse_cir_xlc_hostname(AZ, z_hostname)
                for layer in main_dict['Devices'].keys():
                    for a_index in main_dict['Devices'][layer].keys():
                        for z_index in main_dict['Devices'][layer][a_index]['connections'].keys():
                            if main_dict['Devices'][layer][a_index]['a_hostname'] == a_hostname and \
                                    main_dict['Devices'][layer][a_index]['connections'][z_index]['z_hostname'] == z_hostname:
                                main_dict['Devices'][layer][a_index]['connections'][z_index]['a_interfaces'].append(a_intf)
                                main_dict['Devices'][layer][a_index]['connections'][z_index]['z_interfaces'].append(z_intf)
                for layer in main_dict['Devices'].keys():  # to get BAR->XLC, EDG->BAR, XLC->CIR connections
                    for a_index in main_dict['Devices'][layer].keys():
                        for z_index in main_dict['Devices'][layer][a_index]['connections'].keys():
                            if main_dict['Devices'][layer][a_index]['a_hostname'] == z_hostname and \
                                    main_dict['Devices'][layer][a_index]['connections'][z_index]['z_hostname'] == a_hostname:
                                main_dict['Devices'][layer][a_index]['connections'][z_index]['a_interfaces'].append(z_intf)
                                main_dict['Devices'][layer][a_index]['connections'][z_index]['z_interfaces'].append(a_intf)
            else:
                empty_count += 1
                if empty_count >= 4:
                    break
                empty_count = 0
    return main_dict


def create_edge(main_dict):
    """
    Determines edge rack type and if rack type is dori-s creates info for layer 3 vc-svc device
    :param main_dict:
    :return:
    """
    for a_index in main_dict['Devices']['VC-EDG'].keys():
        main_dict['Devices']['VC-EDG'][a_index]['edge_rack'] = 'vegemite'  # default rack type vegemite unless condition below is met
        for z_index in main_dict['Devices']['VC-EDG'][a_index]['connections'].keys():
            if 'es-svc' in main_dict['Devices']['VC-EDG'][a_index]['connections'][z_index]['z_hostname']:
                main_dict['Devices']['VC-EDG'][a_index]['edge_rack'] = 'dori-s'
                break
        if main_dict['Devices']['VC-EDG'][a_index]['edge_rack'] == 'dori-s':
            a_hostname = main_dict['Devices']['VC-EDG'][a_index]['a_hostname']
            z_index = len(main_dict['Devices']['VC-EDG'][a_index]['connections'].keys()) + 1
            z_hostname = '{}-{}-vc-svc-r{}'.format(AZ, NZ, re.match(r'\w+\d+-vc-edg-r(\d+)', a_hostname).group(1))
            logging.info('Creating connection: {} <> {}.'.format(a_hostname, z_hostname))
            main_dict['Devices']['VC-EDG'][a_index]['connections'][z_index] = OrderedDict()
            main_dict['Devices']['VC-EDG'][a_index]['connections'][z_index]['z_hostname'] = z_hostname
            main_dict['Devices']['VC-EDG'][a_index]['connections'][z_index]['inet_cidr'] = INET_CIDR.get_cidr(a_hostname, z_hostname)


def merge_dict(dct, merge_dct):
    """
    Recursive dict merge. Inspired by :meth:``dict.update()``, instead of
    updating only top-level keys, merge_dict recurses down into dicts nested
    to an arbitrary depth, updating keys. The ``merge_dct`` is merged into ``dct``.
    ref: https://gist.github.com/angstwad/bf22d1822c38a92ec0a9
    :param dct: dict onto which the merge is executed
    :param merge_dct: dct merged into dct
    :return: None
    """
    for key, value in merge_dct.items():
        if (key in dct and isinstance(dct[key], dict)
                and isinstance(merge_dct[key], Mapping)):
            merge_dict(dct[key], merge_dct[key])
        else:
            dct[key] = merge_dct[key]


def fill_serial(main_dict):
    """
    Fill random serial if serial number doesn't exist in mini.yaml
    :param main_dict:
    :return:
    """
    for layer in main_dict['Devices'].keys():  # Create a_hostname dictionary
        for a_index in main_dict['Devices'][layer].keys():
            if not main_dict['Devices'][layer][a_index]['serial']:
                logging.warning('No serial number found for {}, creating random serial'.format(main_dict['Devices'][layer][a_index]['a_hostname']))
                main_dict['Devices'][layer][a_index]['serial'] = random_serial()


def create_yaml(main_dict, args):
    """
    Create yaml file
    :param main_dict:
    :param args:
    :return:
    """
    template_path = '/apollo/env/DXDeploymentTools/templates/'
    # template_path = '/home/abaseer/DXDeploymentTools/src/DXDeploymentTools/configuration/templates/'
    output_path = os.path.join(os.environ['HOME'], 'DXYamls')
    if args.mini is not None:
        logging.info('Reading {}'.format(args.mini))
        template = 'az_yaml_template.jt'
        yaml_name = os.path.join(output_path, '{}.yaml'.format(AZ))
        with open(args.mini) as mini_dict:  # Load mini yaml
            mini_dict = yaml.load(mini_dict)
        merge_dict(main_dict, mini_dict)
        fill_serial(main_dict)
    else:
        template = 'az_mini_yaml_template.jt'
        yaml_name = os.path.join(output_path, '{}-mini.yaml'.format(AZ))

    with open(os.path.join(template_path, template), 'r') as mini_template:
        mini_template = jinja2.Template(mini_template.read())
    logging.info('Creating {}'.format(yaml_name))
    with open(yaml_name, 'w') as yaml_output:
        yaml_output.write(mini_template.render(main_dict))


def main():
    setup_logging()
    args = parse_args()
    validate_input(args)
    logging.info('Reading Cutsheet: {}'.format(args.cutsheet))
    cutsheet = openpyxl.load_workbook('{}'.format(args.cutsheet))
    main_dict = OrderedDict()
    main_dict['Devices'] = OrderedDict()

    hostname_pairs, hostnames_unique = create_hostname_pair(cutsheet)
    main_dict['Region'] = re.match(r'\w\w\w', AZ).group()
    main_dict['Site'] = AZ
    main_dict['NZ'] = NZ

    count_dict = count_hostname(main_dict, hostnames_unique, hostname_pairs)
    create_a_hostname(main_dict, count_dict)
    create_z_hostname(main_dict, count_dict, hostname_pairs)
    create_edge(main_dict)
    create_intf(cutsheet, main_dict)
    create_yaml(main_dict, args)


if __name__ == "__main__":
        main()


