#!/apollo/sbin/envroot "$ENVROOT/python3.6/bin/python3.6"


from dxd_tools_dev.modules import m_yaml
import logging
import argparse

'''
create narg interface string from cutsheet
https://bladerunner.amazon.com/workflows/batch_narg_hostname_interface/versions/prod
'''

import openpyxl
from collections import namedtuple

DEFAULT_COLUMN_DEVICE_A = "A"
DEFAULT_COLUMN_INTERFACE_A = "B"
DEFAULT_COLUMN_DEVICE_Z = "N"
DEFAULT_COLUMN_INTERFACE_Z = "O"
DEFAULT_DEVICE_TYPE = '-br-'

Peer = namedtuple('Peer', ['device_a', 'device_z'])


def _cell_name(row, column_name):
    return column_name + str(row)


class CutSheet:
    def __init__(self, file,
                 column_device_a=DEFAULT_COLUMN_DEVICE_A,
                 column_intf_a=DEFAULT_COLUMN_INTERFACE_A,
                 column_device_z=DEFAULT_COLUMN_DEVICE_Z,
                 column_intf_z=DEFAULT_COLUMN_INTERFACE_Z):
        self._cutsheet_workbook = openpyxl.load_workbook(file, data_only=True)
        self._device_peer = dict()
        self._interfaces_in_device = dict()
        self._devices = set()
        self.set_connections_location(column_device_a=column_device_a,
                                      column_intf_a=column_intf_a,
                                      column_device_z=column_device_z,
                                      column_intf_z=column_intf_z
                                      )

    def get_interest_device_interfaces(self, device_name):
        return self._interfaces_in_device.get(device_name)

    def get_device_list_by_func(self, is_right_device_func):
        result = [device for device in self._devices
                  if is_right_device_func(device)]
        return result

    def get_device_list_by_regex(self, device_type):
        result = [device for device in self._devices
                  if device_type in device]
        return result

    def _add_interfaces_of_device(self, interface_name, device_name):
        if device_name not in self._interfaces_in_device:
            self._interfaces_in_device[device_name] = set()
        self._interfaces_in_device[device_name].add(interface_name)

    def _add_device_peer(self, device_a, interface_a, device_z, interface_z):
        peer = Peer(device_a, device_z)
        if peer not in self._device_peer:
            self._device_peer[peer] = []
        self._device_peer[peer].append([device_a, interface_a, device_z, interface_z])

    def _add_device(self, device_name):
        self._devices.add(device_name)

    def _store_connection(self, device_a, interface_a, device_z, interface_z):
        if not device_a or not interface_a or not device_z or not interface_z:
            return
        device_a = device_a.strip()
        device_z = device_z.strip()
        interface_a = interface_a.strip()
        interface_z = interface_z.strip()

        self._add_device(device_name=device_a)
        self._add_device(device_name=device_z)

        self._add_interfaces_of_device(interface_name=interface_a,
                                       device_name=device_a)
        self._add_interfaces_of_device(interface_name=interface_z,
                                       device_name=device_z)
        self._add_device_peer(device_a, interface_a, device_z, interface_z)

    def set_connections_location(self,
                                 column_device_a,
                                 column_device_z,
                                 column_intf_a,
                                 column_intf_z):
        for sheet in self._cutsheet_workbook:
            records_in_a_device = []
            max_column = sheet.max_column
            max_row = sheet.max_row
            for row in range(1, max_row + 1):
                device_a = sheet[_cell_name(row, column_device_a)].value
                interface_a = sheet[_cell_name(row, column_intf_a)].value
                device_z = sheet[_cell_name(row, column_device_z)].value
                interface_z = sheet[_cell_name(row, column_intf_z)].value
                self._store_connection(device_a, interface_a, device_z, interface_z)


def is_br_device(device_name):
    return '-br-' in device_name


def gen_narg_string(file,
                    column_device_a=DEFAULT_COLUMN_DEVICE_A,
                    column_device_z=DEFAULT_COLUMN_DEVICE_Z,
                    column_intf_a=DEFAULT_COLUMN_INTERFACE_A,
                    column_intf_z=DEFAULT_COLUMN_INTERFACE_Z,
                    device_type=DEFAULT_DEVICE_TYPE):
    cs = CutSheet(file,
                  column_device_a=column_device_a,
                  column_intf_a=column_intf_a,
                  column_device_z=column_device_z,
                  column_intf_z=column_intf_z
                  )
    devices = cs.get_device_list_by_regex(device_type=device_type)
    result = []
    for device in devices:
        intfs = cs.get_interest_device_interfaces(device)
        for intf in intfs:
            a_record = f'{device}_{intf}'
            result.append(a_record)
    return ','.join(result)


def test_cutsheet():
    file = 'BJS20_group3_cutseet_ClaymoreHDv3_v1.2.xlsx'
    r = gen_narg_string(file)
    print(r)


def load_parameter(yaml_file=None):
    if yaml_file is None:
        yaml_file = "console_config.yaml"

    parameters = m_yaml.load_parameter(file=yaml_file)

    # parameters = None

    # with open(yaml_file) as file:
    #    parameters = yaml.load(file, Loader=yaml.FullLoader)
    #    if 'password' in parameters and parameters['password'] == 'input_password':
    #       parameters['password'] = getpass.getpass("pls input your password: ")

    return parameters


def init_logging():
    logger = logging.getLogger()

    logger.setLevel(logging.INFO)

    # make it print to the console.
    console = logging.StreamHandler()
    logger.addHandler(console)

    format_str = '%(asctime)s\t%(levelname)s -- %(processName)s %(filename)s:%(lineno)s -- %(message)s'
    console.setFormatter(logging.Formatter(format_str))


def get_args():
    parser = argparse.ArgumentParser(description='gen device_interface string for batch narg wf')
    parser.add_argument('-y', '--yaml',
                        help='pls give yaml file as input, example in "https://code.amazon.com/packages/DXDeploymentTools/trees/mainline/--/configuration/da/narg/narg_example.yaml"')

    return parser.parse_args()


def main():
    init_logging()
    args = get_args()

    yaml_file = args.yaml
    parameters = load_parameter(yaml_file=yaml_file)

    device_intfs = gen_narg_string(file=parameters.get('cutsheet_file'),
                    column_device_a=parameters.get('column_device_a'),
                    column_intf_a=parameters.get('column_intf_a'),
                    column_device_z=parameters.get('column_device_z'),
                    column_intf_z=parameters.get('column_intf_z'),
                    device_type=parameters.get('device_type')
                    )
    print('To narg device interface from cutsheet, pls use follwing device_interface string in \nhttps://bladerunner.amazon.com/workflows/batch_narg_hostname_interface/versions/prod')
    print(device_intfs)

if __name__ == '__main__':
    from datetime import datetime

    startTime = datetime.now()
    main()
    print(f'script takes {datetime.now() - startTime}')
